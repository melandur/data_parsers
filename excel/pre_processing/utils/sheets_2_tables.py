import os
import time
from collections import defaultdict

import pandas as pd
from loguru import logger
from openpyxl import load_workbook

# from excel.path_master import EXTRACTED_PATH, CASE_WISE_PATH

pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', None)
pd.set_option('display.max_colwidth', None)


class NestedDefaultDict(defaultdict):
    """Nested dict, which can be dynamically expanded on the fly"""

    def __init__(self, *args, **kwargs) -> None:
        super().__init__(NestedDefaultDict, *args, **kwargs)

    def __repr__(self) -> str:
        return repr(dict(self))


class ExtractSheets2Tables:
    def __init__(self, src: str, dst: str, save_intermediate: bool=True) -> None:
        self.src = src
        self.dst = dst
        self.save_intermediate = save_intermediate
        self.tic = time.time()
        self.wb = None
        self.mode = None
        self.count = None
        self.sheet = None
        self.file_path = None
        self.data_name = None
        self.subject_name = None
        self.subject = NestedDefaultDict()
        os.makedirs(self.dst, exist_ok=True)

    def __call__(self) -> None:
        """Extract sheets to tables"""
        for _ in self.loop_files():
            self.get_meta()
            for row in self.loop_row():
                self.detect_table_name(row)
                if self.mode == 'aha_diagram':
                    df = self.extract_aha_diagram(row)
                    self.save(df)
                if self.mode == 'global_roi':
                    df = self.extract_global_roi(row)
                    self.save(df)
                if self.mode == 'aha_polarmap':
                    df = self.extract_aha_polarmap(row)
                    self.save(df)
                if self.mode == 'roi_polarmap':
                    df = self.extract_roi_polarmap(row)
                    self.save(df)
                if self.mode == 'volume':
                    df = self.extract_volume_3d(row)
                    self.save(df)
            logger.info(self.count)

    def __del__(self) -> None:
        """What time is it"""
        logger.info(f'Execution time: {round((time.time() - self.tic) / 60, 1)} minutes')

    def load_file(self, file: str) -> load_workbook:
        """Load file"""
        self.file_path = os.path.join(self.src, file)
        self.subject_name = file.strip('.xlsx')
        return load_workbook(self.file_path, read_only=False, data_only=True, keep_vba=False, keep_links=False)

    def loop_files(self) -> range:
        """Iterate over files"""
        for file in os.listdir(self.src):
            if file.endswith('.xlsx') and not file.startswith('.'):
                logger.info(f'File -> {file}')
                self.wb = self.load_file(file)
                sheet_names = self.wb.sheetnames
                self.sheet = self.wb[sheet_names[0]]
                yield

    def loop_row(self) -> range:
        """Iterate over rows and return certain row numbers"""
        start_row = 229  # start row of the first table
        max_rows = self.sheet.max_row
        self.count = 0
        for row_index in range(start_row, max_rows - 1):
            # considers only left ventricle data
            if 'left' in f'{self.sheet.cell(row=row_index, column=2).value}'.lower():
                yield row_index

    def get_meta(self) -> None:
        """Get meta data from header part"""
        self.subject[self.subject_name] = NestedDefaultDict()
        self.subject[self.subject_name]['meta']['study_date'] = self.sheet.cell(row=212, column=3).value
        self.subject[self.subject_name]['meta']['modality'] = self.sheet.cell(row=219, column=3).value
        self.subject[self.subject_name]['meta']['sequence_name'] = self.sheet.cell(row=223, column=3).value
        self.subject[self.subject_name]['meta']['protocol_name'] = self.sheet.cell(row=224, column=3).value

    def detect_table_name(self, row: int) -> None:
        """Detect table name"""
        data_name = f'{self.sheet.cell(row=row, column=2).value}{self.sheet.cell(row=row, column=3).value}'
        data_name_split = data_name.split('-')
        self.data_name = None
        self.mode = None

        if len(data_name_split) == 3:
            logger.info(data_name_split)
            sub_name_1 = data_name_split[1].replace('Results', '').strip()
            sub_name_1 = sub_name_1.replace(' ', '_').lower()
            sub_name_2 = data_name_split[2].replace('None', '').strip()
            sub_name_2 = sub_name_2.replace('/', '-')
            sub_name_2 = sub_name_2.replace(' ', '_').lower()
            sub_name_2 = sub_name_2.replace('radial', 'radial')
            sub_name_2 = sub_name_2.replace('longitudinal', 'longit')
            sub_name_2 = sub_name_2.replace('circumferential', 'circumf')
            if 'AHA Diagram Data' in data_name_split[0]:
                self.count += 1
                self.data_name = f'aha_{sub_name_1}_{sub_name_2}'
                self.mode = 'aha_diagram'
            elif 'Global and ROI Diagram Data' in data_name_split[0]:
                self.count += 1
                self.data_name = f'global_roi_{sub_name_1}_{sub_name_2}'
                self.mode = 'global_roi'
            elif 'Volume' in data_name_split[2]:
                self.count += 1
                self.data_name = f'volume_{sub_name_1}_(ml)'
                self.mode = 'volume'

        elif len(data_name_split) == 2:
            sub_name_1 = data_name_split[1].replace('Results', '').strip()
            sub_name_1 = sub_name_1.replace(' ', '_').lower()
            if '2D' in data_name_split[1]:
                if 'ROI Polarmap Data' in data_name_split[0] or 'ROI Polarmap Data' in data_name_split[1]:
                    self.count += 1
                    self.data_name = 'roi_polarmap_2d'
                    self.mode = 'roi_polarmap'
                elif 'AHA Polarmap Data' in data_name_split[0]:
                    self.count += 1
                    self.mode = 'aha_polarmap'
                    if 'long' in sub_name_1.lower():
                        self.data_name = 'aha_polarmap_2d_long_axis'
                    elif 'short' in sub_name_1.lower():
                        self.data_name = 'aha_polarmap_2d_short_axis'
                    else:
                        raise ValueError('axis is not defined')

    def _table_row_end_finder(self, start_row: int, column: int, criteria: str or None = None) -> int:
        """Count relative to the start point the number of row until the table ends"""
        for row in range(start_row + 5, start_row + 400, 1):  # 400 is the maximum number of rows to search
            if self.sheet.cell(row=row, column=column).value is criteria:
                return row - start_row - 2  # -2 to account for the header row and the last row
        raise AssertionError(
            f'End of table search range reached, super long table or wrong end criteria -> {start_row}'
        )

    def extract_roi_polarmap(self, row: int) -> pd.DataFrame:
        """Extract roi polarmap"""
        logger.info(f'{row} {self.mode} {self.data_name}')
        row_end = self._table_row_end_finder(row, 2, None)
        df = pd.read_excel(self.file_path, self.subject_name, skiprows=row + 2, nrows=row_end, usecols='B:S')
        df.columns = [
            'slices',
            'roi',
            'peak_strain_radial_%',
            'peak_strain_circumf_%',
            'time_to_peak_1_radial_ms',
            'time_to_peak_1_circumf_ms',
            'peak_systolic_strain_rate_radial_1/s',
            'peak_systolic_strain_rate_circumf_1/s',
            'peak_diastolic_strain_rate_radial_1/s',
            'peak_diastolic_strain_rate_circumf_1/s',
            'peak_displacement_radial_mm',
            'peak_displacement_circumf_mm',
            'time_to_peak_2_radial_ms',
            'time_to_peak_2_circumf_ms',
            'peak_systolic_velocity_radial_mm/s',
            'peak_systolic_velocity_circumf_deg/s',
            'peak_diastolic_velocity_radial_mm/s',
            'peak_diastolic_velocity_circumf_deg/s',
        ]
        return df

    def extract_aha_polarmap(self, row: int) -> pd.DataFrame:
        """Extract aha polarmap"""
        logger.info(f'{row} {self.mode} {self.data_name}')
        row_end = self._table_row_end_finder(row, 2, None)
        df = pd.read_excel(self.file_path, self.subject_name, skiprows=row + 2, nrows=row_end, usecols='B:R')
        if 'short' in self.data_name:
            axis = 'circumf'
            unit = 'deg'
        elif 'long' in self.data_name:
            axis = 'longit'
            unit = 'mm'
        else:
            raise ValueError('axis is not defined')

        df.columns = [
            'aha_segment',
            'peak_strain_radial_%',
            f'peak_strain_{axis}_%',
            'time_to_peak_1_radial_ms',
            f'time_to_peak_1_{axis}_ms',
            'peak_systolic_strain_rate_radial_1/s',
            f'peak_systolic_strain_rate_{axis}_1/s',
            'peak_diastolic_strain_rate_radial_1/s',
            f'peak_diastolic_strain_rate_{axis}_1/s',
            'peak_displacement_radial_mm',
            f'peak_displacement_{axis}_{unit}',
            'time_to_peak_2_radial_ms',
            f'time_to_peak_2_{axis}_ms',
            'peak_systolic_velocity_radial_mm/s',
            f'peak_systolic_velocity_{axis}_{unit}/s',
            'peak_diastolic_velocity_radial_mm/s',
            f'peak_diastolic_velocity_{axis}_{unit}/s',
        ]
        return df

    @staticmethod
    def rearrange_time_helper(df: pd.DataFrame) -> pd.DataFrame or None:
        """Rearrange time columns for 2d"""
        df = df.iloc[:, :-1]  # remove last column
        header = df.columns.tolist()
        counter = 0
        for idx, name in enumerate(header):
            if 'unnamed' in name.lower() or 'ms' in name.lower():
                header[idx] = f'sample_{counter}'
                counter += 1

        df.columns = header
        first_row = df.iloc[0]
        nan_count = first_row.isna().sum()
        first_row = first_row.dropna()
        first_row = first_row.values.tolist()

        if df.iloc[-1].isna().sum() == len(df.columns):
            df = df.drop(df.index[-1])

        for idx, name in enumerate(first_row):
            df.insert(int(idx * 2 + nan_count), f'time_{idx}', name)

        df = df.drop(df.index[0])  # drop the first row
        nan_count = df.isna().sum().sum()
        non_nan_count = df.notna().sum().sum()
        if nan_count < non_nan_count:
            return df
        return None

    @staticmethod
    def rearrange_time_volume(df: pd.DataFrame) -> pd.DataFrame or None:
        """Rearrange time columns for 3d volumes"""
        df = df.iloc[:, :-1]  # remove last column
        header = df.head().columns.values.tolist()
        counter = 0
        for idx, name in enumerate(header):
            if 'unnamed' in name.lower() or 'ms' in name.lower():
                header[idx] = f'sample_{counter}'
                counter += 1

        df.columns = header
        first_row = df.iloc[0]
        nan_count = first_row.isna().sum()
        first_row = first_row.dropna()
        first_row = first_row.values.tolist()

        if df.iloc[-1].isna().sum() == len(df.columns):  # drop the last row if it is empty
            df = df.drop(df.index[-1])

        for idx, name in enumerate(first_row):  # insert time columns
            df.insert(int(idx * 2 + nan_count), f'time_{idx - 1}', name)

        df = df.drop(df.columns[0], axis=1)  # drop the first column
        df = df.drop(df.index[0])  # drop the first row
        nan_count = df.isna().sum().sum()
        non_nan_count = df.notna().sum().sum()
        if nan_count < non_nan_count:
            return df
        return None

    def _last_column_is_empty(self, df: pd.DataFrame) -> bool:
        """Check if the last column is empty"""
        last_column = df.columns[-1]
        if df[last_column].isna().sum() == len(df):
            return True
        return False

    def extract_aha_diagram(self, row: int) -> pd.DataFrame or None:
        """Extract aha diagram 2d"""
        logger.info(f'{row} {self.mode} {self.data_name}')
        row_end = self._table_row_end_finder(row, 2, None)
        df = pd.read_excel(self.file_path, self.subject_name, skiprows=row + 1, nrows=row_end, usecols='B:AB')
        if not df.empty:
            if self._last_column_is_empty(df):
                df = self.rearrange_time_helper(df)
                return df
        logger.warning(f'Empty/invalid dataframe for {self.subject_name} {self.mode} {self.data_name}')
        return None

    def extract_global_roi(self, row: int) -> pd.DataFrame or None:
        """Extract global roi 2d"""
        logger.info(f'{row} {self.mode} {self.data_name}')
        row_end = self._table_row_end_finder(row, 2, None)
        df = pd.read_excel(self.file_path, self.subject_name, skiprows=row + 1, nrows=row_end, usecols='B:AC')
        if not df.empty:
            if self._last_column_is_empty(df):
                df = self.rearrange_time_helper(df)
                return df
        logger.warning(f'Empty/invalid dataframe for {self.subject_name} {self.mode} {self.data_name}')
        return None

    def extract_volume_3d(self, row: int) -> pd.DataFrame or None:
        """Extract volume 3d"""
        logger.info(f'{row} {self.mode} {self.data_name}')
        row_end = self._table_row_end_finder(row, 2, None)
        df = pd.read_excel(self.file_path, self.subject_name, skiprows=row + 1, nrows=row_end, usecols='B:AB')
        if not df.empty:
            if self._last_column_is_empty(df):
                df = self.rearrange_time_volume(df)
                return df
        logger.warning(f'Empty/invalid dataframe for {self.subject_name} {self.mode} {self.data_name}')
        return None

    def save(self, df: pd.DataFrame) -> None:
        """Save dataframe to excel"""
        if df is not None:
            if '2d' in self.data_name:
                d_name = '2d'
            elif '3d' in self.data_name:
                d_name = '3d'
            else:
                raise ValueError(f'Data is not 2d or 3d -> {self.subject_name}')
            file_path = os.path.join(self.dst, self.subject_name, d_name, f'{self.subject_name}_{self.data_name}')
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            df.to_excel(f'{file_path}.xlsx', index=False)


if __name__ == '__main__':
    sheets_2_tables = ExtractSheets2Tables(
        # src=EXTRACTED_PATH,
        # dst=CASE_WISE_PATH,
        src='/home/sebalzer/Documents/Mike_init/tests/train/1_extracted',
        dst='/home/sebalzer/Documents/Mike_init/tests/train/2_case_wise'
    )
    sheets_2_tables()

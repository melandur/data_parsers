import os
import time
from collections import defaultdict

import openpyxl
import pandas as pd
import psutil
from loguru import logger
from openpyxl import load_workbook


class NestedDefaultDict(defaultdict):
    """Nested dict, which can be dynamically expanded on the fly"""

    def __init__(self, *args, **kwargs) -> None:
        super().__init__(NestedDefaultDict, *args, **kwargs)

    def __repr__(self) -> str:
        return repr(dict(self))


class ExtractSheets2Tables:
    def __init__(self, src, dst):
        self.src = src
        self.dst = dst
        self.wb = None
        self.mode = None
        self.count = None
        self.sheet = None
        self.file_path = None
        self.data_name = None
        self.subject_name = None
        self.subject = NestedDefaultDict()
        os.makedirs(self.dst, exist_ok=True)

    def __call__(self):
        """Extract sheets to tables"""
        for _ in self.loop_files():
            self.get_meta()
            for row in self.loop_row():
                self.detect_table_name(row)
                if self.mode == 'aha_diagram':
                    df = self.extract_aha_diagram_2d(row)
                    self.save(df)
                if self.mode == 'global_roi':
                    df = self.extract_global_roi_2d(row)
                    self.save(df)
                if self.mode == 'aha_polarmap':
                    df = self.extract_aha_polarmap_2d(row)
                    self.save(df)
                if self.mode == 'roi_polarmap':
                    df = self.extract_roi_polarmap_2d(row)
                    self.save(df)
            logger.info(self.count)

    def load_file(self, file):
        """Load file"""
        if not file.startswith('.') and file.endswith('.xlsx'):
            self.file_path = os.path.join(self.src, file)
            self.subject_name = file.strip('.xlsx')
            return load_workbook(self.file_path, read_only=False, data_only=True, keep_vba=False, keep_links=False)

    def loop_files(self):
        """Iterate over files"""
        for file in os.listdir(self.src):
            logger.info(f'File -> {file}')
            self.wb = self.load_file(file)
            sheet_names = self.wb.sheetnames
            self.sheet = self.wb[sheet_names[0]]
            yield

    def loop_row(self):
        """Iterate over rows"""
        start_row = 229
        max_rows = self.sheet.max_row
        self.count = 0
        for row_index in range(start_row, max_rows - 1):
            if 'left' in f'{self.sheet.cell(row=row_index, column=2).value}'.lower():
                yield row_index

    def get_meta(self):
        """Get meta data"""
        self.subject[self.subject_name] = NestedDefaultDict()
        self.subject[self.subject_name]['meta']['study_date'] = self.sheet.cell(row=212, column=3).value
        self.subject[self.subject_name]['meta']['modality'] = self.sheet.cell(row=219, column=3).value
        self.subject[self.subject_name]['meta']['sequence_name'] = self.sheet.cell(row=223, column=3).value
        self.subject[self.subject_name]['meta']['protocol_name'] = self.sheet.cell(row=224, column=3).value

    def detect_table_name(self, row):
        """Detect table name"""
        data_name = f'{self.sheet.cell(row=row, column=2).value}{self.sheet.cell(row=row, column=3).value}'
        data_name_split = data_name.split('-')
        self.data_name = None
        self.mode = None

        if len(data_name_split) == 3:
            if '2D' in data_name_split[1]:
                sub_name_1 = data_name_split[1].replace('Results', '').strip()
                sub_name_1 = sub_name_1.replace(' ', '_').lower()
                sub_name_2 = data_name_split[2].replace('None', '').strip()
                sub_name_2 = sub_name_2.replace('/', '-')
                sub_name_2 = sub_name_2.replace(' ', '_').lower()
                sub_name_2 = sub_name_2.replace('radial', 'radial')
                sub_name_2 = sub_name_2.replace('longitudinal', 'longit')
                sub_name_2 = sub_name_2.replace('circumferential', 'circumf')
                if 'AHA Diagram Data' in data_name_split[0]:
                    self.count += 1
                    self.data_name = f'aha_{sub_name_1}_{sub_name_2}'
                    self.mode = 'aha_diagram'
                elif 'Global and ROI Diagram Data' in data_name_split[0]:
                    self.count += 1
                    self.data_name = f'global_roi_{sub_name_1}_{sub_name_2}'
                    self.mode = 'global_roi'

        elif len(data_name_split) == 2:
            sub_name_1 = data_name_split[1].replace('Results', '').strip()
            sub_name_1 = sub_name_1.replace(' ', '_').lower()
            if '2D' in data_name_split[1]:
                if 'ROI Polarmap Data' in data_name_split[0] or 'ROI Polarmap Data' in data_name_split[1]:
                    self.count += 1
                    self.data_name = 'roi_polarmap_2d'
                    self.mode = 'roi_polarmap'
                elif 'AHA Polarmap Data' in data_name_split[0]:
                    self.count += 1
                    self.mode = 'aha_polarmap'
                    if 'long' in sub_name_1.lower():
                        self.data_name = 'aha_polarmap_2d_long_axis'
                    elif 'short' in sub_name_1.lower():
                        self.data_name = 'aha_polarmap_2d_short_axis'
                    else:
                        raise ValueError('axis is not defined')

    def _table_end_finder(self, start_row, column, criteria=None):
        """Count relative to the start point the number of row until the table ends"""
        for row in range(start_row + 5, start_row + 400, 1):  # 400 is the maximum number of rows to search
            if self.sheet.cell(row=row, column=column).value is criteria:
                return row - start_row - 2  # -2 to account for the header row and the last row
        raise AssertionError(
            f'End of table search range reached, super long table or wrong end criteria -> {start_row}'
        )

    def extract_roi_polarmap_2d(self, row):
        """Extract roi polarmap 2d"""
        logger.info(f'{row} {self.mode} {self.data_name}')
        row_end = self._table_end_finder(row, 2, None)
        df = pd.read_excel(self.file_path, self.subject_name, skiprows=row + 2, nrows=row_end, usecols='B:S')
        df.columns = [
            'slices',
            'roi',
            'peak_strain_radial_%',
            'peak_strain_circumf_%',
            'time_to_peak_1_radial_ms',
            'time_to_peak_1_circumf_ms',
            'peak_systolic_strain_rate_radial_1/s',
            'peak_systolic_strain_rate_circumf_1/s',
            'peak_diastolic_strain_rate_radial_1/s',
            'peak_diastolic_strain_rate_circumf_1/s',
            'peak_displacement_radial_mm',
            'peak_displacement_circumf_mm',
            'time_to_peak_2_radial_ms',
            'time_to_peak_2_circumf_ms',
            'peak_systolic_velocity_radial_mm/s',
            'peak_systolic_velocity_circumf_deg/s',
            'peak_diastolic_velocity_radial_mm/s',
            'peak_diastolic_velocity_circumf_deg/s',
        ]
        return df

    def extract_aha_polarmap_2d(self, row):
        """Extract aha polarmap 2d"""
        logger.info(f'{row} {self.mode} {self.data_name}')
        row_end = self._table_end_finder(row, 2, None)
        df = pd.read_excel(self.file_path, self.subject_name, skiprows=row + 2, nrows=row_end, usecols='B:R')
        if 'short' in self.data_name:
            axis = 'circumf'
            unit = 'deg'
        elif 'long' in self.data_name:
            axis = 'longit'
            unit = 'mm'
        else:
            raise ValueError('axis is not defined')

        df.columns = [
            'aha_segment',
            'peak_strain_radial_%',
            f'peak_strain_{axis}_%',
            'time_to_peak_1_radial_ms',
            f'time_to_peak_1_{axis}_ms',
            'peak_systolic_strain_rate_radial_1/s',
            f'peak_systolic_strain_rate_{axis}_1/s',
            'peak_diastolic_strain_rate_radial_1/s',
            f'peak_diastolic_strain_rate_{axis}_1/s',
            'peak_displacement_radial_mm',
            f'peak_displacement_{axis}_{unit}',
            'time_to_peak_2_radial_ms',
            f'time_to_peak_2_{axis}_ms',
            'peak_systolic_velocity_radial_mm/s',
            f'peak_systolic_velocity_{axis}_{unit}/s',
            'peak_diastolic_velocity_radial_mm/s',
            f'peak_diastolic_velocity_{axis}_{unit}/s',
        ]
        return df

    def extract_aha_diagram_2d(self, row):
        """Extract aha diagram 2d"""
        logger.info(f'{row} {self.mode} {self.data_name}')
        row_end = self._table_end_finder(row, 2, None)
        df = pd.read_excel(self.file_path, self.subject_name, skiprows=row + 2, nrows=row_end, usecols='B:AA')
        header = df.head()
        header = header.columns.values.tolist()
        header = [x for x in header if type(x) != str]
        header = [f'time_{x}_ms' for x in header]
        header[:0] = ['aha_segment']
        if len(header) == len(df.columns):
            df.columns = header
            return df
        return None

    def extract_global_roi_2d(self, row):
        """Extract global roi 2d"""
        logger.info(f'{row} {self.mode} {self.data_name}')
        row_end = self._table_end_finder(row, 2, None)
        df = pd.read_excel(self.file_path, self.subject_name, skiprows=row + 2, nrows=row_end, usecols='B:AB')
        header = df.head()
        header = header.columns.values.tolist()
        header = [x for x in header if type(x) != str]
        header = [f'time_{x}_ms' for x in header]
        header[:0] = ['slice', 'roi']
        if len(header) == len(df.columns):
            df.columns = header
            return df
        return None

    def save(self, df):
        """Save dataframe to excel"""
        if df is not None:
            file_path = os.path.join(self.dst, self.subject_name, f'{self.subject_name}_{self.data_name}.xlsx')
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            df.to_excel(file_path, index=False)  # index=False to avoid the index column in the excel file


if __name__ == '__main__':
    x = time.time()

    sheets_2_tables = ExtractSheets2Tables(
        src='/home/melandur/Data/extracted',
        dst='/home/melandur/Downloads/tables_without_index',
    )
    sheets_2_tables()

    logger.info(f'Execution time: {round((time.time() - x))/60} minutes')


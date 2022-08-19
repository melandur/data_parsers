import os
import time

import openpyxl
import psutil
from loguru import logger
from openpyxl import load_workbook


class ExtractWorkbook2Sheets:
    def __init__(self, src_file, dst_folder):
        self.src_file = src_file
        self.dst_folder = dst_folder
        self.wb = None
        self.sheet = None
        self.subject_name = None
        os.makedirs(self.dst_folder, exist_ok=True)

    def __call__(self):
        """Extract workbook to sheets"""
        total_memory = psutil.virtual_memory().total / (1024**3)  # in GB
        if total_memory < 30:  # warn if memory is less than 32GB
            raise MemoryError('Not enough memory to extract workbook to sheets, 32 GB of RAM is required')
        logger.info(f'Extract workbook to sheets is running...')

        if not self.src_file.endswith('.xlsx') and not os.path.exists(self.src_file):
            raise ValueError(f'{self.src_file} is not a valid ".xlsx" file')

        self.extract_sheets()

    def extract_sheets(self):
        """Extract sheets"""
        wb = self.load_file()  # load workbook
        for sheet_name in wb.sheetnames:  # loop through sheets
            if '_' in sheet_name or '2 ' in sheet_name:  # skip sheets with '_' or '2 ' in name
                if '#' not in sheet_name:  # skip sheets with '#' in name
                    logger.info(f'Extract -> {sheet_name}')
                    extracted_sheet = wb[sheet_name]  # extract sheet
                    new_wb = openpyxl.Workbook()  # create new workbook
                    single_sheet = new_wb.active  # get active sheet
                    if '_' in sheet_name:
                        new_sheet_name = sheet_name.split('_')[1]
                    else:
                        new_sheet_name = sheet_name[2:]
                    single_sheet.title = new_sheet_name
                    for row in extracted_sheet:  # copy sheet to new workbook
                        for cell in row:
                            single_sheet[cell.coordinate].value = cell.value

                    new_wb.save(f'{os.path.join(self.dst_folder, new_sheet_name)}.xlsx')
                    new_wb.close()
                    del new_wb
                    del single_sheet
        del wb

    def load_file(self):
        """Load file"""
        if not self.src_file.startswith('.'):
            self.subject_name = self.src_file.strip('.xlsx')
            return load_workbook(self.src_file, read_only=False, data_only=True, keep_vba=False, keep_links=False)


if __name__ == '__main__':
    x = time.time()

    workbook_2_sheets = ExtractWorkbook2Sheets(
        src_file='/home/melandur/Data/Myocarditis/CRF_control/CRF_controls/CRF FlamBer/D. Strain_v3b_FlamBer_61-120.xlsx',
        dst_folder='/home/melandur/Downloads/hello/',
    )
    workbook_2_sheets()

    logger.info(f'Execution time: {round((time.time() - x))/60} minutes')

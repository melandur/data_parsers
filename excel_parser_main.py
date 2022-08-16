import os
import pickle
from openpyxl import load_workbook
from collections import defaultdict
from loguru import logger
import pandas as pd


class NestedDefaultDict(defaultdict):
    """Nested dict, which can be dynamically expanded on the fly"""

    def __init__(self, *args, **kwargs) -> None:
        super().__init__(NestedDefaultDict, *args, **kwargs)

    def __repr__(self) -> str:
        return repr(dict(self))


class ExcelParser:

    def __init__(self, src, dst):
        """"""
        self.src = src
        self.dst = dst
        self.wb = None
        self.mode = None
        self.sheet = None
        self.file_path = None
        self.data_name = None
        self.subject_name = None
        self.subject = NestedDefaultDict()
        os.makedirs(self.dst, exist_ok=True)

    def __call__(self):
        """"""
        for _ in self.loop_files():
            self.get_meta()
            for row in self.loop_row():
                self.extract_name(row)
                if self.mode == 'aha_diagram':
                    self.extract_aha_diagram_2d(row)
                if self.mode == 'global_roi':
                    self.extract_global_roi_2d(row)
                if self.mode == 'aha_polarmap':
                    self.extract_aha_polarmap_2d(row)
                if self.mode == 'roi_polarmap':
                    self.extract_roi_polarmap_2d(row)
            logger.info(self.count)

    def load_file(self, file, extension='pkl'):
        """"""
        if file.endswith(extension) and not file.startswith('.'):
            self.file_path = os.path.join(self.src, file)
            if extension == 'pkl':
                with open(self.file_path, 'rb') as file_object:
                    return pickle.load(file_object)
            self.subject_name = file.strip('.xlsx')
            return load_workbook(self.file_path,
                                 read_only=False,
                                 data_only=True,
                                 keep_vba=False,
                                 keep_links=False)

    def extract_sheets_as_xlsx(self):
        """"""
        # print(os.listdir(self.src))
        # for file in os.listdir(self.src):
        file = self.src
        wb = self.load_file(file, 'xlsx')
        for sheet_name in wb.sheetnames:
            if '_' in sheet_name or '2 ' in sheet_name:
                if '#' not in sheet_name:
                    logger.info(f'Extract -> {sheet_name}')
                    extracted_sheet = wb[sheet_name]
                    new_wb = openpyxl.Workbook()
                    single_sheet = new_wb.active
                    if '_' in sheet_name:
                        new_sheet_name = sheet_name.split('_')[1]
                    else:
                        new_sheet_name = sheet_name[2:]
                    single_sheet.title = new_sheet_name
                    for row in extracted_sheet:
                        for cell in row:
                            single_sheet[cell.coordinate].value = cell.value
                    new_wb.save(f'{os.path.join(self.dst, new_sheet_name)}.xlsx')
                    new_wb.close()
                    del new_wb
                    del single_sheet
        del wb

    def loop_files(self):
        """Iterate over files"""
        for file in os.listdir(self.src):
            logger.info(f'File -> {file}')
            self.wb = self.load_file(file, extension='xlsx')
            sheet_names = self.wb.sheetnames
            self.sheet = self.wb[sheet_names[0]]
            yield

    def loop_row(self):
        """"""
        start_row = 229
        max_rows = self.sheet.max_row
        self.count = 0
        for row_index in range(start_row, max_rows - 1):
            if 'left' in f'{self.sheet.cell(row=row_index, column=2).value}'.lower():
                yield row_index

    def get_meta(self):
        self.subject[self.subject_name] = NestedDefaultDict()
        self.subject[self.subject_name]['meta']['study_date'] = self.sheet.cell(row=212, column=3).value
        self.subject[self.subject_name]['meta']['modality'] = self.sheet.cell(row=219, column=3).value
        self.subject[self.subject_name]['meta']['sequence_name'] = self.sheet.cell(row=223, column=3).value
        self.subject[self.subject_name]['meta']['protocol_name'] = self.sheet.cell(row=224, column=3).value

    def extract_name(self, row):
        data_name = f'{self.sheet.cell(row=row, column=2).value}{self.sheet.cell(row=row, column=3).value}'
        data_name_split = data_name.split('-')
        self.data_name = None
        self.mode = None

        if len(data_name_split) == 3:
            if '2D' in data_name_split[1]:
                sub_name_1 = data_name_split[1].replace('Results', '').strip()
                sub_name_1 = sub_name_1.replace(' ', '_').lower()
                sub_name_2 = data_name_split[2].replace('None', '').strip()
                sub_name_2 = sub_name_2.replace('/', '-')
                sub_name_2 = sub_name_2.replace(' ', '_').lower()
                sub_name_2 = sub_name_2.replace('radial', 'rad')
                sub_name_2 = sub_name_2.replace('longitudinal', 'lon')
                sub_name_2 = sub_name_2.replace('circumferential', 'cir')
                if 'AHA Diagram Data' in data_name_split[0]:
                    self.count += 1
                    self.data_name = f'aha_{sub_name_1}_{sub_name_2}'
                    self.mode = 'aha_diagram'
                elif 'Global and ROI Diagram Data' in data_name_split[0]:
                    self.count += 1
                    self.data_name = f'global_roi_{sub_name_1}_{sub_name_2}'
                    self.mode = 'global_roi'

        elif len(data_name_split) == 2:
            sub_name_1 = data_name_split[1].replace('Results', '').strip()
            sub_name_1 = sub_name_1.replace(' ', '_').lower()
            if '2D' in data_name_split[1]:
                if 'ROI Polarmap Data' in data_name_split[0] or 'ROI Polarmap Data' in data_name_split[1]:
                    self.count += 1
                    self.data_name = 'roi_polarmap_2d'
                    self.mode = 'roi_polarmap'
                elif 'AHA Polarmap Data' in data_name_split[0]:
                    self.count += 1
                    self.mode = 'aha_polarmap'
                    if 'long' in sub_name_1.lower():
                        self.data_name = 'aha_polarmap_2d_long_axis'
                    elif 'short' in sub_name_1.lower():
                        self.data_name = 'aha_polarmap_2d_short_axis'
                    else:
                        raise ValueError('axis is not defined')

    def _table_end_finder(self, start_row, column=2, criteria=None):
        """Count relative to the start point the number of row until the table ends"""
        for row in range(start_row + 5, start_row + 400, 1):
            if self.sheet.cell(row=row, column=column).value is criteria:
                return row - start_row - 2
        raise AssertionError(f'End of table search range reached, super long table or wrong end criteria -> {start_row}')

    def extract_roi_polarmap_2d(self, row):
        """Extract roi polarmap 2d"""
        logger.info(f'{row} {self.mode} {self.data_name}')
        row_end = self._table_end_finder(row, 2, None)
        df = pd.read_excel(self.file_path, self.subject_name, skiprows=row + 2, nrows=row_end, usecols='B:M')
        df.columns = ['slices', 'roi', 'peak_strain_rad_%', 'peak_strain_cir_%', 'time_to_peak_rad_ms',
                      'time_to_peak_cir_ms', 'peak_systolic_strain_rate_rad_1/s', 'peak_systolic_strain_rate_cir_1/s',
                      'peak_diastolic_strain_rate_rad_1/s', 'peak_diastolic_strain_rate_cir_1/s',
                      'peak_displacement_rad_mm', 'peak_displacement_cir_mm']
        folder_path = os.path.join(self.dst, self.subject_name)
        os.makedirs(folder_path, exist_ok=True)
        df.to_excel(os.path.join(folder_path, f'{self.data_name}.xlsx'))

    def extract_aha_polarmap_2d(self, row):
        """Extract aha polarmap 2d"""
        logger.info(f'{row} {self.mode} {self.data_name}')
        row_end = self._table_end_finder(row, 2, None)
        df = pd.read_excel(self.file_path, self.subject_name, skiprows=row + 2, nrows=row_end, usecols='B:L')
        df.columns = ['aha_segment', 'peak_strain_rad_%', 'peak_strain_cir_%', 'time_to_peak_rad_ms',
                      'time_to_peak_cir_ms', 'peak_systolic_strain_rate_rad_1/s', 'peak_systolic_strain_rate_cir_1/s',
                      'peak_diastolic_strain_rate_rad_1/s', 'peak_diastolic_strain_rate_cir_1/s',
                      'peak_displacement_rad_mm', 'peak_displacement_cir_mm']
        folder_path = os.path.join(self.dst, self.subject_name)
        os.makedirs(folder_path, exist_ok=True)
        df.to_excel(os.path.join(folder_path, f'{self.data_name}.xlsx'))

    def extract_aha_diagram_2d(self, row):
        """Extract aha diagram 2d"""
        logger.info(f'{row} {self.mode} {self.data_name}')
        row_end = self._table_end_finder(row, 2, None)
        df = pd.read_excel(self.file_path, self.subject_name, skiprows=row + 2, nrows=row_end, usecols='B:AA')
        header = [f'time_{x}_ms' for x in range(25)]
        header[:0] = ['aha_segment']
        df.columns = header
        folder_path = os.path.join(self.dst, self.subject_name)
        os.makedirs(folder_path, exist_ok=True)
        df.to_excel(os.path.join(folder_path, f'{self.data_name}.xlsx'))

    def extract_global_roi_2d(self, row):
        """Extract global roi 2d"""
        logger.info(f'{row} {self.mode} {self.data_name}')
        row_end = self._table_end_finder(row, 2, None)
        df = pd.read_excel(self.file_path, self.subject_name, skiprows=row + 2, nrows=row_end, usecols='B:AB')
        header = [f'time_{x}_ms' for x in range(25)]
        header[:0] = ['slice', 'roi']
        df.columns = header
        folder_path = os.path.join(self.dst, self.subject_name)
        os.makedirs(folder_path, exist_ok=True)
        df.to_excel(os.path.join(folder_path, f'{self.data_name}.xlsx'))


if __name__ == '__main__':
    import time

    x = time.time()

    ep = ExcelParser(
        src='/home/melandur/Data/Myocarditis/raw_csv/part_4',
        dst='/home/melandur/Data/Myocarditis/test_csv_processing')

    ep()

    logger.info(round((time.time() - x)))
